"""
report_generator.py
--------------------
Generates the exportable deliverables:
  - Recruitment_Report.xlsx (multi-sheet, formatted with OpenPyXL)
  - Recruitment_Report.pdf  (executive summary + charts)
  - Recruitment_KPIs.csv

Author: Recruitment Analytics Team
"""

from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from src.utils import Paths, get_logger

logger = get_logger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(bold=True, size=14, name="Calibri", color="1F4E78")


class ReportGenerator:
    """Builds the Excel, PDF, and CSV recruitment reports."""

    def __init__(self, df: pd.DataFrame, kpi_summary: pd.DataFrame) -> None:
        self.df = df
        self.kpi_summary = kpi_summary
        Paths.REPORTS_DIR.mkdir(exist_ok=True)

    def _write_sheet(self, ws, dataframe: pd.DataFrame, title: str) -> None:
        ws["A1"] = title
        ws["A1"].font = TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(dataframe.columns), 1))

        for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center")

        n_cols = max(len(dataframe.columns), 1)
        for col_idx in range(1, n_cols + 1):
            letter = ws.cell(row=3, column=col_idx).column_letter
            length = max(
                (len(str(ws.cell(row=r, column=col_idx).value)) for r in range(3, ws.max_row + 1)
                 if ws.cell(row=r, column=col_idx).value is not None),
                default=0,
            )
            ws.column_dimensions[letter].width = min(max(length + 2, 12), 40)

    def generate_excel_report(self, filename: str = "Recruitment_Report.xlsx") -> Path:
        """Build a multi-sheet, formatted Excel workbook."""
        wb = Workbook()

        ws_kpi = wb.active
        ws_kpi.title = "Executive KPIs"
        self._write_sheet(ws_kpi, self.kpi_summary, "Recruitment KPI Summary")

        ws_role = wb.create_sheet("By Job Role")
        role_df = self.df["Job Roles"].value_counts().reset_index()
        role_df.columns = ["Job Role", "Applications"]
        self._write_sheet(ws_role, role_df, "Applications by Job Role")

        chart = BarChart()
        chart.title = "Applications by Job Role"
        data_ref = Reference(ws_role, min_col=2, min_row=3, max_row=3 + min(len(role_df), 15))
        cats_ref = Reference(ws_role, min_col=1, min_row=4, max_row=3 + min(len(role_df), 15))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws_role.add_chart(chart, "E3")

        ws_dept = wb.create_sheet("Department Hiring")
        dept_df = self.df.groupby("Department").agg(
            Applications=("Job Applicant Name", "count"),
            Hires=("Application_Status", lambda s: (s == "Hired").sum()),
        ).reset_index()
        self._write_sheet(ws_dept, dept_df, "Department Hiring Overview")

        ws_recruiter = wb.create_sheet("Recruiter Performance")
        rec_df = self.df.groupby("Recruiter").agg(
            Applications=("Job Applicant Name", "count"),
            Hires=("Application_Status", lambda s: (s == "Hired").sum()),
        ).reset_index()
        self._write_sheet(ws_recruiter, rec_df, "Recruiter Performance")

        ws_source = wb.create_sheet("Hiring Source")
        src_df = self.df["Hiring_Source"].value_counts().reset_index()
        src_df.columns = ["Hiring Source", "Applications"]
        self._write_sheet(ws_source, src_df, "Applications by Hiring Source")

        ws_raw = wb.create_sheet("Cleaned Data (Sample)")
        self._write_sheet(ws_raw, self.df.head(500), "Cleaned Candidate Data (first 500 rows)")

        out_path = Paths.REPORTS_DIR / filename
        wb.save(out_path)
        logger.info("Excel report saved: %s", out_path)
        return out_path

    def generate_pdf_report(self, filename: str = "Recruitment_Report.pdf") -> Path:
        """Build a PDF executive summary with an embedded KPI table and key charts."""
        out_path = Paths.REPORTS_DIR / filename

        with PdfPages(out_path) as pdf:
            # Page 1: KPI summary table
            fig, ax = plt.subplots(figsize=(8.5, 11))
            ax.axis("off")
            ax.set_title("Recruitment Analytics — Executive Report", fontsize=16, fontweight="bold", pad=20)
            table_data = self.kpi_summary.values.tolist()
            table = ax.table(
                cellText=table_data,
                colLabels=self.kpi_summary.columns.tolist(),
                loc="center",
                cellLoc="left",
            )
            table.auto_set_font_size(False)
            table.set_fontsize(9)
            table.scale(1, 1.6)
            pdf.savefig(fig)
            plt.close(fig)

            # Page 2: Applications by job role
            fig, ax = plt.subplots(figsize=(8.5, 6))
            self.df["Job Roles"].value_counts().head(10).plot(kind="barh", ax=ax, color="#2E86AB")
            ax.set_title("Top 10 Job Roles by Applications")
            ax.invert_yaxis()
            plt.tight_layout()
            pdf.savefig(fig)
            plt.close(fig)

            # Page 3: Department hiring
            fig, ax = plt.subplots(figsize=(8.5, 6))
            dept_hires = self.df[self.df["Application_Status"] == "Hired"]["Department"].value_counts()
            dept_hires.plot(kind="bar", ax=ax, color="#F18F01")
            ax.set_title("Hires by Department")
            plt.xticks(rotation=45, ha="right")
            plt.tight_layout()
            pdf.savefig(fig)
            plt.close(fig)

        logger.info("PDF report saved: %s", out_path)
        return out_path

    def generate_kpi_csv(self, filename: str = "Recruitment_KPIs.csv") -> Path:
        out_path = Paths.REPORTS_DIR / filename
        self.kpi_summary.to_csv(out_path, index=False)
        logger.info("KPI CSV saved: %s", out_path)
        return out_path

    def generate_all(self) -> None:
        self.generate_excel_report()
        self.generate_pdf_report()
        self.generate_kpi_csv()
        logger.info("Reports Generated | Excel, PDF, and CSV outputs complete")
